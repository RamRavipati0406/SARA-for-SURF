import openpyxl
from openpyxl import Workbook
from langchain_openai import ChatOpenAI
from langchain.prompts import ChatPromptTemplate, HumanMessagePromptTemplate
from langchain.chains import LLMChain
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.pydantic_v1 import BaseModel, Field
import json
from tqdm import tqdm
import time
import langchain_core.exceptions


# Define your desired data structure.
class JobRequirements(BaseModel):
    Soft_Skills: str = Field(description="Soft skills required for the job")
    Experience: str = Field(description="Experience required for the job")
    Education: str = Field(description="Minimum Educational Qualifications")
    SED_Knowledge: str = Field(description="Systems engineering domain knowledge familiarity required for the job")
    Modelling_Skills_Langugages: str = Field(description="Modelling languages/skills the candidate is required for the job")
    tools: str = Field(description="Tools and technologies required for the job")

# Set up a parser + inject instructions into the prompt template.
parser = JsonOutputParser(pydantic_object=JobRequirements)

def create_agent(print_thinking=False):
    try:
        # LLM
        llm = ChatOpenAI(model_name='#your-prefered-GPT-Model', temperature=0,
                         openai_api_key='#your-GPT-API-key',
                         # tiktoken_model_name="gpt-3.5-turbo-16k"
                         )

        # Prompt
        prompt = ChatPromptTemplate(
            messages=[
                HumanMessagePromptTemplate.from_template("{question}")
            ],
            partial_variables={"format_instructions": parser.get_format_instructions()},
        )

        conversation_agent = LLMChain(
            llm=llm,
            prompt=prompt,
            verbose=print_thinking,
            output_parser=parser
        )
        return conversation_agent

    except Exception as ex:
        print(ex)


if __name__ == '__main__':

    description_list = []
    # Load the Excel file
    workbook = openpyxl.load_workbook('THEBOOK-SURF.xlsx')
    new_wb = Workbook()
    new_sheet = new_wb.active

    # Select the worksheet by name
    worksheet = workbook.active

    new_headers = ['Soft Skills', 'Experience Required', 'Education Requirements', 
               'SE Domain Knowledge', 'Modeliing Languages/Skills', 'Tool and Technologies']
    
    col_num = 1
    for header in new_headers:
        new_sheet.cell(row=1,column=col_num, value=header)
        col_num = col_num+1

    for row in worksheet.iter_rows(min_row=2,
                                max_row=worksheet.max_row, 
                                min_col=6, 
                                max_col=6):
        for cell in row:
            if cell.value is None:
                break
            else:
                description_list.append(cell.value)
    
    agent = create_agent()
    results = []

    print('\n')
    for desc in tqdm(description_list):
        prompt = f'''
                For the provided job description, tell me
                1. Soft skills required for the job
                2. Experience required for the job
                3. Minimum Educational Qualifications
                4. Systems engineering domain knowledge familiarity required for the job
                5. what modelling languages/skills the candidate is requirwed for the job
                6. tools and technologies required for the job
                if you cannot figure it out, reply 'Unknown'
                I want output in JSON and use the following keys:
                'Soft Skills', 'Experience', 'Education', 'SED Knowledge', 'Modelling Skills/Langugages', 'tools'
                    I do not want sentences, give me comma seperated words if multiple, else a single word

                JOB DESCRIPTION = {desc} '''
        try:    
            response = agent.invoke({"question": prompt})
            json_data = response['text']
        except langchain_core.exceptions.OutputParserException:
            json_data = '{"Soft Skills": "Unknown", "Experience": "Unknown", "Education": "Unknown", "SED Knowledge": "Unknown", "Modelling Skills/Langugages": "Unknown", "tools": "Unknown"}'
        print(json_data)
        results.append(json_data)
        time.sleep(1)
    
    for result in results:
        new_sheet.append([value for value in result.values()])
        new_wb.save('4turbo--THEBOOK-SURF.xlsx')

    # print(type(json_data))
    # print(json_data)
    # print('Soft Skills: ',json_data['Soft Skills'])
    # print('Experience: ',json_data['Experience'])
    print('\nWriting Data to New Excel File')
    workbook.close()
    new_wb.close()
